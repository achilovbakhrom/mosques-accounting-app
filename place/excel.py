# views.py
from drf_spectacular.utils import extend_schema
from rest_framework.exceptions import PermissionDenied
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView
from rest_framework.parsers import MultiPartParser
from rest_framework.response import Response
from rest_framework import status
from openpyxl import load_workbook

from core.models import Place


class UploadExcelView(APIView):
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser]

    @extend_schema(
        request={
            'multipart/form-data': {
                'type': 'object',
                'properties': {
                    'file': {
                        'type': 'string',
                        'format': 'binary',
                        'description': 'Excel file containing place data.',
                    },
                },
                'required': ['file'],
            }
        },
        responses={
            201: {'type': 'object', 'properties': {'status': {'type': 'string'}}},
            400: {'type': 'object', 'properties': {'error': {'type': 'string'}}},
            500: {'type': 'object', 'properties': {'error': {'type': 'string'}}},
        },
        description='Upload an Excel file to populate the database with hierarchical place data.',
        tags=['Place Upload'],
    )
    def post(self, request):

        current_user = request.user
        if current_user.role != 'admin':
            raise PermissionDenied("Only admin has access to this endpoint.")

        file_obj = request.FILES.get('file', None)
        if not file_obj:
            return Response({'error': 'No file provided.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            # Load the workbook
            workbook = load_workbook(filename=file_obj)
            # Select the second sheet (index starts from 0)
            sheet = workbook.worksheets[1]

            parent_place = None

            # Iterate over the rows starting from C4
            for row in sheet.iter_rows(min_row=4, min_col=3, max_col=4, values_only=True):
                place_name, inn = row
                if place_name:
                    if not inn:
                        # This is a parent place
                        parent_place = Place.objects.create(name=place_name, is_mosque=False)
                    else:
                        # This is a child place
                        Place.objects.create(name=place_name, inn=inn, parent=parent_place, is_mosque=True)
                else:
                    # Skip rows without a place name
                    continue

            return Response({'status': 'Data uploaded successfully.'}, status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
